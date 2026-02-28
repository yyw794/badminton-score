#!/usr/bin/env python3
"""
Excel 对阵表转 Web JSON 格式
将 lineup_scheduler.py 生成的 Excel 文件转换为 Web 应用可用的 data.json 格式
并同时保存到 SQLite 数据库
"""

import openpyxl
import json
import os
from datetime import datetime
import sys

# 导入数据库模块
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from db import init_db, save_event_data


def parse_excel_to_json(excel_path: str, output_path: str, event_name: str = None):
    """解析 Excel 对阵表并生成 JSON 文件。
    
    Args:
        excel_path: Excel 文件路径
        output_path: JSON 输出路径
        event_name: 赛事名称，默认为文件名
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # 从文件名提取赛事名称
    if event_name is None:
        event_name = os.path.basename(excel_path).replace('.xlsx', '')
    
    # 解析配置信息（第 2 行）
    config_text = ws["A2"].value or ""
    court_count = 3
    if "场地数：" in config_text:
        try:
            court_count = int(config_text.split("场地数：")[1].split("个")[0])
        except:
            pass
    
    # 解析比赛数据（从第 4 行开始）
    matches = []
    match_id = 1
    
    for row in range(4, ws.max_row + 1):
        round_num = ws.cell(row=row, column=1).value
        court = ws.cell(row=row, column=2).value
        match_type = ws.cell(row=row, column=3).value
        team_a_str = ws.cell(row=row, column=4).value
        team_b_str = ws.cell(row=row, column=7).value
        
        if not round_num or not match_type:
            continue
        
        # 解析场地号（如 "1 号" -> 1）
        if isinstance(court, str) and "号" in court:
            court = int(court.replace("号", ""))
        
        # 解析队伍（如 "张三/李四" -> ["张三", "李四"]）
        team_a = team_a_str.split("/") if team_a_str else []
        team_b = team_b_str.split("/") if team_b_str else []
        
        match = {
            "id": f"m{match_id}",
            "round": round_num,
            "court": court,
            "type": match_type,
            "teamA": team_a,
            "teamB": team_b,
            "scoreA": [0, 0],
            "scoreB": [0, 0],
            "status": "pending"
        }
        matches.append(match)
        match_id += 1
    
    # 计算球员统计
    player_stats = {}
    for match in matches:
        match_type = match["type"]
        for player in match["teamA"] + match["teamB"]:
            if player not in player_stats:
                player_stats[player] = {"total": 0, "男双": 0, "女双": 0, "混双": 0}
            player_stats[player]["total"] += 1
            player_stats[player][match_type] += 1
    
    # 构建输出数据
    output_data = {
        "eventName": event_name,
        "courtCount": court_count,
        "matches": matches,
        "playerStats": player_stats,
        "exportTime": datetime.now().isoformat()
    }
    
    # 写入 JSON 文件
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    print(f"✓ 已导出 {len(matches)} 场比赛到 {output_path}")
    print(f"  赛事名称：{event_name}")
    print(f"  场地数量：{court_count}")
    print(f"  球员数量：{len(player_stats)}")
    
    return output_data


def main():
    """主函数：从默认路径读取 Excel 并导出 JSON。"""
    # 默认路径：docs 目录的上一级是项目根目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)  # 项目根目录
    excel_path = os.path.join(project_dir, "排阵/对阵表.xlsx")
    output_path = os.path.join(script_dir, "data.json")
    
    # 从命令行参数获取路径和活动名称
    event_name = None
    if len(sys.argv) > 1:
        if sys.argv[1].startswith("-"):
            # 跳过选项参数
            pass
        else:
            event_name = sys.argv[1]
    if len(sys.argv) > 2:
        output_path = sys.argv[2]
    
    # 如果没有传入活动名称，从微信接龙文件提取
    if event_name is None:
        signup_file = os.path.join(project_dir, "排阵/微信接龙.txt")
        if os.path.exists(signup_file):
            with open(signup_file, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#"):
                        # 尝试提取日期
                        import re
                        # 支持格式：2026-02-28, 2 月 28 日，02-28
                        match = re.search(r'(\d{4}-\d{2}-\d{2}|\d{1,2} 月\d{1,2} 日|\d{1,2}-\d{1,2})', line)
                        if match:
                            date_str = match.group(1)
                            # 标准化日期格式
                            if '月' in date_str and '日' in date_str:
                                parts = date_str.replace('月', '-').replace('日', '').split('-')
                                date_str = f"2026-{parts[0].zfill(2)}-{parts[1].zfill(2)}"
                            elif '-' in date_str and len(date_str) == 5:
                                parts = date_str.split('-')
                                date_str = f"2026-{parts[0].zfill(2)}-{parts[1].zfill(2)}"
                            event_name = f"{date_str} 活动"
                        else:
                            event_name = "羽毛球训练赛"
                        break
    
    if event_name is None:
        event_name = "羽毛球训练赛"
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误：找不到 Excel 文件 {excel_path}")
        print("  请先运行：python lineup_scheduler.py")
        return
    
    output_data = parse_excel_to_json(excel_path, output_path, event_name)
    
    # 保存到数据库
    print("\n💾 保存到数据库...")
    init_db()
    save_event_data(event_name, output_data["matches"], output_data["courtCount"])
    
    print(f"\n使用方法:")
    print(f"  1. 打开 docs/index.html 即可使用 Web 记分功能")
    print(f"  2. 数据已自动更新到 docs/data.json")
    print(f"  3. 历史数据已保存到 data.db")


if __name__ == "__main__":
    main()
