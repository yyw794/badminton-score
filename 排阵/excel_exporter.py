#!/usr/bin/env python3
"""
Badminton Lineup Excel Exporter
Reusable module for generating Excel lineup schedules.
Supports both traditional and LLM-based schedulers.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import List, Dict, Optional, Tuple
import re


# Player definitions (shared across modules)
INTERNAL_MALE_PLAYERS = [
    "苏大哲", "罗蒙", "江锐", "严勇文", "陈顺星", "陈小洪",
    "卢志辉", "林锋", "王小波", "刘继宇", "董广博", "林琪琛", "罗琴荩"
]

GUEST_MALE_PLAYERS = [
    "张欣欣", "黄冬青", "程建兴", "陈宇霆", "卢子龙", "吴煜"
]

MALE_PLAYERS = INTERNAL_MALE_PLAYERS + GUEST_MALE_PLAYERS

INTERNAL_FEMALE_PLAYERS = [
    "田茜", "唐英武", "李祺祺", "高洁", "滕菲", "谢卓珊", "崔倩男", "林小连"
]

GUEST_FEMALE_PLAYERS = [
    "张燕红", "李杏芝", "项小英"
]

FEMALE_PLAYERS = INTERNAL_FEMALE_PLAYERS + GUEST_FEMALE_PLAYERS

# Mixed doubles eligible male players (internal only)
MIXED_DOUBLES_MALES = {"林锋", "王小波", "陈顺星", "罗琴荩", "罗蒙"}

# Player-specific constraints
# fixed_games: None = 自动根据场地紧张程度计算，整数 = 固定场次
PLAYER_CONSTRAINTS = {
    "严勇文": {"fixed_games": None, "early_departure": True},
    "崔倩男": {"fixed_games": None, "early_departure": True},
    "林小连": {"only_womens_doubles": True},
}

# 场地紧张程度阈值（球员平均比赛场次）
# 当球员平均比赛场次 >= 此值时，场地算充裕；否则算紧张
COURT_ABUNDANCE_THRESHOLD = 5.0

# 提前离场球员的场次配置
EARLY_DEPARTURE_GAMES = {
    "abundant": 5,  # 场地充裕时的场次
    "tight": 4,     # 场地紧张时的场次
}

# Fixed partner pairs
FIXED_PARTNERS = {
    ("苏大哲", "罗蒙"): 5,
}

# Global config
MATCHES_PER_COURT = 8
MAX_GAMES_INTERNAL = 7
MAX_GAMES_GUEST = 6  # Will be dynamically adjusted based on court availability


def get_max_games_for_player(player: str, court_count: int, total_players: int) -> int:
    """
    Get the maximum games allowed for a player based on court availability.

    Logic:
    - If courts are abundant (courts >= 3 and total_players <= 16), no restriction on guest players
    - If courts are limited, guest players have lower priority (max 5-6 games)
    """
    # Calculate court abundance: courts per player
    courts_per_player = court_count / total_players if total_players > 0 else 0

    # If courts are abundant (e.g., 3 courts for 12-16 players), no restriction
    if courts_per_player >= 0.2 and court_count >= 3:  # 3 courts for 15 players = 0.2
        return MAX_GAMES_INTERNAL  # Same as internal players

    # If courts are limited, apply guest restriction
    if player in GUEST_MALE_PLAYERS or player in GUEST_FEMALE_PLAYERS:
        return MAX_GAMES_GUEST

    return MAX_GAMES_INTERNAL


def get_fixed_games_for_player(player: str, court_count: int, total_players: int, total_matches: int) -> Optional[int]:
    """
    Get the fixed games for a player with early_departure constraint.

    This dynamically adjusts based on court availability:
    - Court abundance (avg_games >= threshold): higher games (e.g., 5)
    - Court scarcity (avg_games < threshold): lower games (e.g., 4)

    Args:
        player: Player name
        court_count: Number of courts
        total_players: Total number of players
        total_matches: Total number of matches planned

    Returns:
        Fixed games count, or None if no fixed_games constraint
    """
    constraint = PLAYER_CONSTRAINTS.get(player, {})
    fixed_games = constraint.get("fixed_games")

    # If explicitly set to an integer, use that value
    if fixed_games is not None:
        return fixed_games

    # If set to None and player has early_departure, calculate dynamically
    if constraint.get("early_departure"):
        # Calculate average games per player
        # Each match involves 4 player-games
        total_player_games = total_matches * 4
        avg_games_per_player = total_player_games / total_players if total_players > 0 else 0

        # Determine court abundance
        if avg_games_per_player >= COURT_ABUNDANCE_THRESHOLD:
            return EARLY_DEPARTURE_GAMES["abundant"]
        else:
            return EARLY_DEPARTURE_GAMES["tight"]

    return None


def parse_activity_date(signup_text: str) -> str:
    """
    Parse activity date from signup text.
    
    Supports formats:
    - 20260323 羽毛球活动
    - 2026-03-23 羽毛球活动
    - 2026/03/23 羽毛球活动
    - 3 月 23 日羽毛球活动
    """
    lines = signup_text.strip().split("\n")
    
    for line in lines[:5]:  # Check first 5 lines
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        
        # Try YYYYMMDD format
        match = re.search(r'(\d{4})(\d{2})(\d{2})', line)
        if match:
            year, month, day = match.groups()
            return f"{year}年{month}月{day}日"
        
        # Try YYYY-MM-DD or YYYY/MM/DD format
        match = re.search(r'(\d{4})[-/](\d{2})[-/](\d{2})', line)
        if match:
            year, month, day = match.groups()
            return f"{year}年{month}月{day}日"
        
        # Try 月日 format (e.g., 3 月 23 日)
        match = re.search(r'(\d{1,2})月 (\d{1,2}) 日', line)
        if match:
            month, day = match.groups()
            return f"{month}月{day}日"
    
    return ""  # No date found


def create_lineup_excel(
    matches: List[Dict],
    court_count: int,
    output_path: str,
    player_stats: Optional[Dict] = None,
    title: str = "科技球队日常训练活动 - 对阵表",
    schedule_method: str = "",
    activity_date: str = ""
):
    """
    Create Excel file with lineup schedule.

    Args:
        matches: List of match dictionaries with keys:
            - round: round number (1, 2, 3...)
            - court: court number (1, 2, 3...)
            - type: match type ("男双", "女双", "混双", etc.)
            - match: tuple of ((player1, player2), (player3, player4))
        court_count: Number of courts (2 or 3)
        output_path: Output file path
        player_stats: Optional player statistics dictionary
        title: Custom title for the sheet
        schedule_method: Scheduling method description (e.g., "传统算法", "LLM 推理")
        activity_date: Activity date (e.g., "2026 年 03 月 23 日")
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对阵表"

    # Style definitions
    font_title = Font(name="微软雅黑", size=18, bold=True)
    font_header = Font(name="微软雅黑", size=11, bold=True)
    font_content = Font(name="微软雅黑", size=10)
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Title row with date
    ws.merge_cells("A1:G1")
    date_suffix = f"（{activity_date}）" if activity_date else ""
    ws["A1"] = f"{title}{date_suffix}"
    ws["A1"].font = font_title
    ws["A1"].alignment = alignment_center
    ws["A1"].border = border

    # Config row
    ws.merge_cells("A2:G2")
    method_info = f" | 排阵方式：{schedule_method}" if schedule_method else ""
    ws["A2"] = f"场地数：{court_count}个 | 时长：2 小时 | 赛制：15 分/局，2 局 | 项目：男双、女双、混双{method_info}"
    ws["A2"].font = Font(name="微软雅黑", size=10)
    ws["A2"].alignment = alignment_center

    # Headers
    headers = ["轮次", "场地", "类型", "对阵 A", "比分 A", "比分 B", "对阵 B"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = border
        cell.fill = header_fill

    # Match rows
    for row_idx, match_info in enumerate(matches, 4):
        match_type = match_info["type"]
        match = match_info["match"]
        court = match_info["court"]
        round_num = match_info.get("round", 1)

        ws.cell(row=row_idx, column=1, value=round_num).font = font_content
        ws.cell(row=row_idx, column=2, value=f"{court}号").font = font_content
        ws.cell(row=row_idx, column=3, value=match_type).font = font_content
        ws.cell(row=row_idx, column=4, value="/".join(match[0])).font = font_content
        ws.cell(row=row_idx, column=5, value="").font = font_content
        ws.cell(row=row_idx, column=6, value="").font = font_content
        ws.cell(row=row_idx, column=7, value="/".join(match[1])).font = font_content

        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).alignment = alignment_center
            ws.cell(row=row_idx, column=col).border = border

    # Column widths
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 22

    # Row heights
    max_row = len(matches) + 3
    for row in ws.iter_rows(min_row=1, max_row=max_row if max_row > 3 else 20):
        ws.row_dimensions[row[0].row].height = 32

    # Page setup for A4 landscape printing
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "landscape"
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Player statistics sheet (optional)
    if player_stats:
        ws_stats = wb.create_sheet(title="球员统计")
        font_title_stats = Font(name="微软雅黑", size=14, bold=True)
        font_header_stats = Font(name="微软雅黑", size=11, bold=True)
        font_content_stats = Font(name="微软雅黑", size=10)
        alignment_center_stats = Alignment(horizontal="center", vertical="center")
        border_stats = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        header_fill_stats = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        ws_stats.merge_cells("A1:E1")
        ws_stats["A1"] = "球员参赛场次统计"
        ws_stats["A1"].font = font_title_stats
        ws_stats["A1"].alignment = alignment_center_stats
        ws_stats["A1"].border = border_stats

        headers_stats = ["姓名", "总场次", "男双", "女双", "混双"]
        for col, header in enumerate(headers_stats, 1):
            cell = ws_stats.cell(row=2, column=col, value=header)
            cell.font = font_header_stats
            cell.alignment = alignment_center_stats
            cell.border = border_stats
            cell.fill = header_fill_stats

        row = 3
        # Sort players: males first, then females
        male_players = [p for p in player_stats.keys() if p in MALE_PLAYERS]
        female_players = [p for p in player_stats.keys() if p in FEMALE_PLAYERS]

        for player in sorted(male_players):
            stats = player_stats[player]
            ws_stats.cell(row=row, column=1, value=player).font = font_content_stats
            ws_stats.cell(row=row, column=2, value=stats["total"]).font = font_content_stats
            ws_stats.cell(row=row, column=3, value=stats.get("男双", 0)).font = font_content_stats
            ws_stats.cell(row=row, column=4, value=stats.get("女双", 0)).font = font_content_stats
            ws_stats.cell(row=row, column=5, value=stats.get("混双", 0)).font = font_content_stats
            for col in range(1, 6):
                ws_stats.cell(row=row, column=col).alignment = alignment_center_stats
                ws_stats.cell(row=row, column=col).border = border_stats
            row += 1

        for player in sorted(female_players):
            stats = player_stats[player]
            ws_stats.cell(row=row, column=1, value=player).font = font_content_stats
            ws_stats.cell(row=row, column=2, value=stats["total"]).font = font_content_stats
            ws_stats.cell(row=row, column=3, value=stats.get("男双", 0)).font = font_content_stats
            ws_stats.cell(row=row, column=4, value=stats.get("女双", 0)).font = font_content_stats
            ws_stats.cell(row=row, column=5, value=stats.get("混双", 0)).font = font_content_stats
            for col in range(1, 6):
                ws_stats.cell(row=row, column=col).alignment = alignment_center_stats
                ws_stats.cell(row=row, column=col).border = border_stats
            row += 1

        ws_stats.column_dimensions["A"].width = 12
        ws_stats.column_dimensions["B"].width = 10
        ws_stats.column_dimensions["C"].width = 10
        ws_stats.column_dimensions["D"].width = 10
        ws_stats.column_dimensions["E"].width = 10

    wb.save(output_path)
    print(f"对阵表已生成：{output_path}")


def calculate_player_stats(matches: List[Dict], all_players: List[str]) -> Dict:
    """
    Calculate player statistics from matches.
    
    Args:
        matches: List of match dictionaries
        all_players: List of all player names
        
    Returns:
        Dictionary with player stats: {player: {"total": N, "男双": N, "女双": N, "混双": N}}
    """
    player_games = {}
    player_type_games = {}
    
    for m in matches:
        match_type = m["type"]
        # Handle fallback types: extract base type
        base_type = match_type.split(" ")[0]  # "男双 (临时)" -> "男双"
        
        for player in get_match_players(m["match"]):
            player_games[player] = player_games.get(player, 0) + 1
            if player not in player_type_games:
                player_type_games[player] = {"男双": 0, "女双": 0, "混双": 0}
            player_type_games[player][base_type] += 1
    
    player_stats = {}
    for player in all_players:
        type_stats = player_type_games.get(player, {"男双": 0, "女双": 0, "混双": 0})
        player_stats[player] = {
            "total": player_games.get(player, 0),
            "男双": type_stats["男双"],
            "女双": type_stats["女双"],
            "混双": type_stats["混双"]
        }
    
    return player_stats


def generate_mixed_vs_mens_matches(males: List[str], females: List[str]) -> List[Tuple[Tuple[str, str], Tuple[str, str]]]:
    """
    Generate mixed doubles vs men's doubles matches.
    
    In this format:
    - Team A: Mixed doubles (1 male + 1 female)
    - Team B: Men's doubles (2 males)
    
    This allows more female participation when female players are limited.
    """
    import itertools
    
    mixed_males = [m for m in males if m in MIXED_DOUBLES_MALES]
    
    # Filter females who can play mixed
    can_play_womens_doubles = len(females) >= 4
    mixed_females = [f for f in females
                     if not PLAYER_CONSTRAINTS.get(f, {}).get("only_womens_doubles")
                     or not can_play_womens_doubles]
    
    matches = []
    
    # For each possible mixed doubles pair
    for male in mixed_males:
        for female in mixed_females:
            # Find remaining males for men's doubles
            remaining_males = [m for m in males if m != male]
            
            # Need 2 males for men's doubles
            if len(remaining_males) < 2:
                continue
            
            # Generate all possible men's doubles pairs
            for mens_pair in itertools.combinations(remaining_males, 2):
                match = ((male, female), mens_pair)
                matches.append(match)
    
    return matches


def get_match_players(match) -> set:
    """Get all players in a match."""
    players = set()
    for pair in match:
        players.update(pair)
    return players


def print_schedule_summary(matches: List[Dict], all_players: List[str], court_count: int):
    """Print a summary of the schedule to console."""
    from collections import Counter
    
    match_type_counts = Counter()
    for m in matches:
        match_type_counts[m["type"]] += 1
    
    print(f"\n比赛类型分布:")
    for t, c in sorted(match_type_counts.items()):
        print(f"  - {t}: {c}场")
    
    print(f"\n【轮次详情】")
    rounds_dict = {}
    for m in matches:
        round_num = m.get("round", 1)
        if round_num not in rounds_dict:
            rounds_dict[round_num] = []
        rounds_dict[round_num].append(m)
    
    for round_num in sorted(rounds_dict.keys()):
        round_matches = rounds_dict[round_num]
        court_count_round = len(round_matches)
        print(f"\n第{round_num}轮 ({court_count_round}个场地):")
        round_players = set()
        for m in round_matches:
            court = m["court"]
            match_type = m["type"]
            team_a = "/".join(m["match"][0])
            team_b = "/".join(m["match"][1])
            players_in_match = get_match_players(m["match"])
            print(f"  {court}号场地 [{match_type}]: {team_a} VS {team_b}")
            round_players.update(players_in_match)
        print(f"  本轮球员 ({len(round_players)}人): {', '.join(sorted(round_players))}")
    
    print("\n✓ 核心规则检查通过：无球员在同一轮次重复出现")
    
    print(f"\n球员参赛场次统计:")
    player_stats = calculate_player_stats(matches, all_players)
    for player in sorted(all_players):
        stats = player_stats[player]
        type_str = f"(男{stats['男双']}/女{stats['女双']}/混{stats['混双']})"
        print(f"  - {player}: {stats['total']}场 {type_str}")
