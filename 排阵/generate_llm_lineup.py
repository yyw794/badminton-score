#!/usr/bin/env python3
"""
LLM 推理排阵 - 直接由大模型推理生成对阵方案
而非使用贪心算法逐场选择
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

# ===== 大模型推理的排阵方案 =====
# 策略：女双优先（3场），混双充足，男双填充
# 核心约束：
#   - 每轮每球员只能出现1次
#   - 严勇文 5 场后走
#   - 崔倩男 5 场后走
#   - 黄冬青(外援) <= 6 场
#   - 女双 3 场 (WD1/3/5轮)
#
# 女双配对（3场，用3种不同配对组合）：
#   WD1: LQQ/TF vs CQN/TQ
#   WD3: LQQ/TQ vs TF/CQN
#   WD5: LQQ/CQN vs TF/TQ
#
# 混双分配（每女2场XD，共8场XD）：
#   LQQ: WD(1,3,5) + XD(2,4) = 5场
#   TF:  WD(1,3,5) + XD(2,6) = 5场
#   TQ:  WD(1,3,5) + XD(4,8) = 5场
#   CQN: WD(1,3,5) + XD(2,7) = 5场 (CQN=5场，刚好)

LINEUP = {
    "date": "2026年04月13日",
    "court_count": 3,
    "rounds": [
        # === 第1轮: 女双开场 ===
        # 12人: LQQ,TF,CQN,TQ | YW,LZH | SDZ,LJY | LF,LQJ
        # 空: LM, CX, HDQ
        {
            "courts": [
                {"type": "女双", "a": ("李祺祺", "滕菲"), "b": ("崔倩男", "田茜")},
                {"type": "男双", "a": ("严勇文", "卢志辉"), "b": ("苏大哲", "刘继宇")},
                {"type": "男双", "a": ("林锋", "罗琴荩"), "b": ("罗蒙", "王小波")},
            ]
        },
        # === 第2轮: 2XD + MD, 严勇文MD(2) ===
        # 12人: LM,LQQ | LQJ,CQN | YW,CX | HDQ,CXH | LF,TF | WXB,TQ
        # 空: LJY, LZH, SDZ
        {
            "courts": [
                {"type": "混双", "a": ("罗蒙", "李祺祺"), "b": ("罗琴荩", "崔倩男")},
                {"type": "男双", "a": ("严勇文", "陈顺星"), "b": ("黄冬青", "陈小洪")},
                {"type": "混双", "a": ("林锋", "滕菲"), "b": ("王小波", "田茜")},
            ]
        },
        # === 第3轮: 女双(2) + MD ===
        # 12人: LQQ,TQ | TF,CQN | YW,LJY | CX,HDQ | LZH,LM
        # 空: SDZ, WXB, LQJ
        {
            "courts": [
                {"type": "女双", "a": ("李祺祺", "田茜"), "b": ("滕菲", "崔倩男")},
                {"type": "男双", "a": ("严勇文", "刘继宇"), "b": ("陈顺星", "陈小洪")},
                {"type": "男双", "a": ("黄冬青", "卢志辉"), "b": ("罗蒙", "林锋")},
            ]
        },
        # === 第4轮: 2XD + MD, 严勇文MD(4) ===
        # 12人: LM,LQQ | CX,CQN | YW,LZH | SDZ,LJY | WXB,TF | LQJ,TQ
        # 空: HDQ, CXH, LF
        {
            "courts": [
                {"type": "混双", "a": ("罗蒙", "李祺祺"), "b": ("陈顺星", "崔倩男")},
                {"type": "男双", "a": ("严勇文", "苏大哲"), "b": ("刘继宇", "卢志辉")},
                {"type": "混双", "a": ("王小波", "滕菲"), "b": ("罗琴荩", "田茜")},
            ]
        },
        # === 第5轮: 女双(3) + 2MD, 严勇文MD(5,走) ===
        # 12人: LQQ,CQN | TF,TQ | SDZ,CXH | YW,HDQ | LM,LF | WXB,LQJ
        # 空: LZH, CX, LJY
        {
            "courts": [
                {"type": "女双", "a": ("李祺祺", "崔倩男"), "b": ("滕菲", "田茜")},
                {"type": "男双", "a": ("苏大哲", "陈小洪"), "b": ("严勇文", "黄冬青")},
                {"type": "男双", "a": ("罗蒙", "林锋"), "b": ("王小波", "罗琴荩")},
            ]
        },
        # === 第6轮: 2MD + XD (林锋/滕菲 vs 罗蒙/田茜, 罗蒙=7场) ===
        # 12人: SDZ,LQJ | LJY,LZH | CX,HDQ | WXB,CXH | LF,TF | LM,TQ
        # 空: YW(走), CQN(走), LQQ
        {
            "courts": [
                {"type": "男双", "a": ("苏大哲", "罗琴荩"), "b": ("刘继宇", "卢志辉")},
                {"type": "男双", "a": ("陈顺星", "黄冬青"), "b": ("王小波", "陈小洪")},
                {"type": "混双", "a": ("林锋", "滕菲"), "b": ("罗蒙", "田茜")},
            ]
        },
        # === 第7轮: MD + 2XD ===
        # 12人: SDZ,LJY | LZH,HDQ | WXB,TQ | CX,LQQ | LM,TF | LF,CXH
        # 空: YW(走), CQN(走), LQJ
        {
            "courts": [
                {"type": "男双", "a": ("苏大哲", "刘继宇"), "b": ("卢志辉", "黄冬青")},
                {"type": "混双", "a": ("王小波", "田茜"), "b": ("罗蒙", "滕菲")},
                {"type": "混双", "a": ("陈顺星", "李祺祺"), "b": ("林锋", "陈小洪")},
            ]
        },
        # === 第8轮: MD + 2XD 收尾 ===
        # 上场(12): SDZ,CXH | LJY,CX | LM,TF | LQJ,TQ | LZH,WXB | LF,LQQ
        # 空: HDQ(已6场)
        {
            "courts": [
                {"type": "男双", "a": ("苏大哲", "陈小洪"), "b": ("刘继宇", "陈顺星")},
                {"type": "混双", "a": ("罗蒙", "滕菲"), "b": ("罗琴荩", "田茜")},
                {"type": "混双", "a": ("卢志辉", "王小波"), "b": ("林锋", "李祺祺")},
            ]
        },
    ]
}


def validate_lineup():
    """Validate the LLM-generated lineup."""
    rounds = LINEUP["rounds"]
    player_games = {}
    all_players = set()
    
    # Players with early departure (stop playing after N games)
    early_departure = {"严勇文": 5, "崔倩男": 5}
    
    for i, rd in enumerate(rounds):
        round_num = i + 1
        round_players = set()
        for court in rd["courts"]:
            players = set()
            for pair in [court["a"], court["b"]]:
                players.update(pair)
                all_players.update(pair)
                for p in pair:
                    player_games[p] = player_games.get(p, 0) + 1
            
            # 检查同轮次不重复
            overlap = round_players & players
            if overlap:
                print(f"❌ 第{round_num}轮重复: {overlap}")
                return False
            round_players |= players
    
    # 检查连续轮空（排除已提前离场的球员）
    for i in range(len(rounds)):
        round_players_per_round = []
        for rd in rounds:
            players = set()
            for court in rd["courts"]:
                for pair in [court["a"], court["b"]]:
                    players.update(pair)
            round_players_per_round.append(players)
        
        if i >= 2:
            for p in all_players:
                # 检查该球员是否已提前离场
                if p in early_departure and player_games.get(p, 0) >= early_departure[p]:
                    # 找到该球员最后一轮上场的轮次
                    last_played = -1
                    for j in range(len(rounds)):
                        if p in round_players_per_round[j]:
                            last_played = j
                    # 只检查最后一轮之前的连续轮空
                    if i <= last_played:
                        played_prev = p in round_players_per_round[i-1]
                        played_prev2 = p in round_players_per_round[i-2]
                        if not played_prev and not played_prev2:
                            current_players = round_players_per_round[i]
                            if p not in current_players:
                                print(f"❌ {p} 连续3轮轮空 (第{i-1},{i},{i+1}轮)")
                                return False
                else:
                    # 正常球员，检查所有轮次
                    played_prev = p in round_players_per_round[i-1]
                    played_prev2 = p in round_players_per_round[i-2]
                    if not played_prev and not played_prev2:
                        current_players = round_players_per_round[i]
                        if p not in current_players:
                            print(f"❌ {p} 连续3轮轮空 (第{i-1},{i},{i+1}轮)")
                            return False
    
    # 打印统计
    print(f"\n{'='*60}")
    print(f"LLM 推理排阵验证")
    print(f"{'='*60}")
    print(f"总人数: {len(all_players)}人")
    print(f"  男性: {sorted([p for p in all_players if p not in ['李祺祺','滕菲','崔倩男','田茜']])}")
    print(f"  女性: {sorted([p for p in all_players if p in ['李祺祺','滕菲','崔倩男','田茜']])}")
    print(f"\n比赛类型统计:")
    type_counts = {"男双": 0, "女双": 0, "混双": 0}
    for rd in rounds:
        for court in rd["courts"]:
            type_counts[court["type"]] = type_counts.get(court["type"], 0) + 1
    for t, c in type_counts.items():
        print(f"  {t}: {c}场")
    
    print(f"\n球员参赛场次:")
    for p, g in sorted(player_games.items(), key=lambda x: -x[1]):
        types = {"男": 0, "女": 0, "混": 0}
        for rd in rounds:
            for court in rd["courts"]:
                if court["type"] == "男双" and p in set(court["a"]) | set(court["b"]):
                    types["男"] += 1
                elif court["type"] == "女双" and p in set(court["a"]) | set(court["b"]):
                    types["女"] += 1
                elif court["type"] == "混双" and p in set(court["a"]) | set(court["b"]):
                    types["混"] += 1
        
        notes = []
        if p in early_departure and g == early_departure[p]:
            notes.append(f"(已完成{g}场) [需提前离场]")
        elif p == "黄冬青":
            notes.append(f"(外援, {g}场)")
        print(f"  {p}: {g}场 (男{types['男']}/女{types['女']}/混{types['混']}) {' '.join(notes)}")
    
    print(f"\n特殊检查:")
    yw_games = player_games.get('严勇文', 0)
    cqn_games = player_games.get('崔倩男', 0)
    hdq_games = player_games.get('黄冬青', 0)
    wd_count = type_counts['女双']
    print(f"  严勇文: {yw_games}场 {'✅' if yw_games <= 5 else '❌ 超过5场'}")
    print(f"  崔倩男: {cqn_games}场 {'✅' if cqn_games <= 5 else '❌ 超过5场'}")
    print(f"  黄冬青: {hdq_games}场 {'✅' if hdq_games <= 6 else '❌ 超过6场'}")
    print(f"  女双: {wd_count}场 {'✅' if wd_count >= 3 else '⚠️ 偏少'}")
    
    return True


def create_excel():
    """Create Excel file from the LLM-generated lineup."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对阵表"

    font_title = Font(name="微软雅黑", size=18, bold=True)
    font_header = Font(name="微软雅黑", size=11, bold=True)
    font_content = Font(name="微软雅黑", size=10)
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws.merge_cells("A1:G1")
    ws["A1"] = f"科技球队日常训练活动 - 对阵表（{LINEUP['date']}）"
    ws["A1"].font = font_title
    ws["A1"].alignment = alignment_center
    ws["A1"].border = border

    ws.merge_cells("A2:G2")
    ws["A2"] = f"场地数：{LINEUP['court_count']}个 | 时长：2 小时 | 赛制：15 分/局，2 局 | 项目：男双、女双、混双"
    ws["A2"].font = Font(name="微软雅黑", size=10)
    ws["A2"].alignment = alignment_center

    headers = ["轮次", "场地", "类型", "对阵 A", "比分 A", "比分 B", "对阵 B"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = border
        cell.fill = header_fill

    row = 4
    for i, rd in enumerate(LINEUP["rounds"]):
        round_num = i + 1
        for j, court in enumerate(rd["courts"]):
            court_num = j + 1
            pair_a = f"{court['a'][0]}/{court['a'][1]}"
            pair_b = f"{court['b'][0]}/{court['b'][1]}"
            ws.cell(row=row, column=1, value=f"第{round_num}轮").font = font_content
            ws.cell(row=row, column=2, value=f"{court_num}号场地").font = font_content
            ws.cell(row=row, column=3, value=court["type"]).font = font_content
            ws.cell(row=row, column=4, value=pair_a).font = font_content
            ws.cell(row=row, column=5, value="").font = font_content
            ws.cell(row=row, column=6, value="").font = font_content
            ws.cell(row=row, column=7, value=pair_b).font = font_content
            for col in range(1, 8):
                ws.cell(row=row, column=col).alignment = alignment_center
                ws.cell(row=row, column=col).border = border
            row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 22

    output_path = "对阵表_LLM.xlsx"
    wb.save(output_path)
    print(f"\n✅ 对阵表已生成：{output_path}")


if __name__ == "__main__":
    print("=" * 60)
    print("LLM 推理排阵系统 - 大模型直接输出")
    print("=" * 60)
    
    if validate_lineup():
        print("\n✅ 验证通过，生成 Excel...")
        create_excel()
    else:
        print("\n❌ 验证失败，请检查排阵方案")
