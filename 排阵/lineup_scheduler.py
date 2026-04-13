#!/usr/bin/env python3
"""
Badminton Lineup Scheduler
Generates balanced matchups for mixed doubles, men's doubles, and women's doubles.
Single sheet output optimized for A4 printing.
Supports player-specific constraints (fixed games, early departure) and fixed partners.
Supports guest players (former employees) with lower priority than internal employees.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import itertools
from typing import List, Tuple, Dict, Optional
import random
import re


# Player definitions
INTERNAL_MALE_PLAYERS = [
    "苏大哲", "罗蒙", "江锐", "严勇文", "陈顺星", "陈小洪",
    "卢志辉", "林锋", "王小波", "刘继宇", "董广博", "林琪琛", "罗琴荩"
]

GUEST_MALE_PLAYERS = [
    "张欣欣", "黄冬青", "程建兴", "陈宇霆", "卢子龙", "吴煜"
]

MALE_PLAYERS = INTERNAL_MALE_PLAYERS + GUEST_MALE_PLAYERS

INTERNAL_FEMALE_PLAYERS = [
    "田茜", "唐英武", "李祺祺", "高洁", "滕菲", "谢卓珊", "崔倩男", "林小连"
]

GUEST_FEMALE_PLAYERS = [
    "张燕红", "李杏芝", "项小英"
]

FEMALE_PLAYERS = INTERNAL_FEMALE_PLAYERS + GUEST_FEMALE_PLAYERS


def parse_activity_date(signup_text: str) -> str:
    """
    Parse activity date from signup text.
    
    Supports formats:
    - 20260323 羽毛球活动
    - 2026-03-23 羽毛球活动
    - 2026/03/23 羽毛球活动
    - 3 月 23 日羽毛球活动
    """
    lines = signup_text.strip().split("\n")
    
    for line in lines[:5]:  # Check first 5 lines
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        
        # Try YYYYMMDD format
        match = re.search(r'(\d{4})(\d{2})(\d{2})', line)
        if match:
            year, month, day = match.groups()
            return f"{year}年{month}月{day}日"
        
        # Try YYYY-MM-DD or YYYY/MM/DD format
        match = re.search(r'(\d{4})[-/](\d{2})[-/](\d{2})', line)
        if match:
            year, month, day = match.groups()
            return f"{year}年{month}月{day}日"
        
        # Try 月日 format (e.g., 3 月 23 日)
        match = re.search(r'(\d{1,2})月 (\d{1,2}) 日', line)
        if match:
            month, day = match.groups()
            return f"{month}月{day}日"
    
    return ""  # No date found

# Mixed doubles eligible male players (internal only)
MIXED_DOUBLES_MALES = {"林锋", "王小波", "陈顺星", "罗琴荩", "罗蒙"}

# Player-specific constraints
# fixed_games: None = 自动根据场地紧张程度计算，整数 = 固定场次
PLAYER_CONSTRAINTS = {
    "严勇文": {"fixed_games": None, "early_departure": True},
    "崔倩男": {"fixed_games": None, "early_departure": True},
    "林小连": {"only_womens_doubles": True},  # 只打女双
}

# 场地紧张程度阈值（球员平均比赛场次）
# 当球员平均比赛场次 >= 此值时，场地算充裕；否则算紧张
COURT_ABUNDANCE_THRESHOLD = 5.0

# 提前离场球员的场次配置
EARLY_DEPARTURE_GAMES = {
    "abundant": 5,  # 场地充裕时的场次
    "tight": 4,     # 场地紧张时的场次
}

# Fixed partner pairs
FIXED_PARTNERS = {
}

# Global config
MATCHES_PER_COURT = 8
MAX_GAMES_INTERNAL = 7  # Internal employees max games
MAX_GAMES_GUEST = 6     # Guest players max games (lower priority)


def is_internal_player(player: str) -> bool:
    """Check if a player is an internal employee."""
    return player in INTERNAL_MALE_PLAYERS or player in INTERNAL_FEMALE_PLAYERS


def is_guest_player(player: str) -> bool:
    """Check if a player is a guest player (former employee)."""
    return player in GUEST_MALE_PLAYERS or player in GUEST_FEMALE_PLAYERS


def get_max_games_for_player(player: str) -> int:
    """Get the maximum games allowed for a player based on their status."""
    if is_guest_player(player):
        return MAX_GAMES_GUEST
    return MAX_GAMES_INTERNAL


def get_fixed_games_for_player(player: str, total_matches: int, total_players: int) -> Optional[int]:
    """
    Get the fixed games for a player with early_departure constraint.

    This dynamically adjusts based on court availability:
    - Court abundance (avg_games >= threshold): higher games (e.g., 5)
    - Court scarcity (avg_games < threshold): lower games (e.g., 4)

    Args:
        player: Player name
        total_matches: Total number of matches planned
        total_players: Total number of players

    Returns:
        Fixed games count, or None if no fixed_games constraint
    """
    constraint = PLAYER_CONSTRAINTS.get(player, {})
    fixed_games = constraint.get("fixed_games")

    # If explicitly set to an integer, use that value
    if fixed_games is not None:
        return fixed_games

    # If set to None and player has early_departure, calculate dynamically
    if constraint.get("early_departure"):
        # Calculate average games per player
        # Each match involves 4 player-games
        total_player_games = total_matches * 4
        avg_games_per_player = total_player_games / total_players if total_players > 0 else 0

        # Determine court abundance
        if avg_games_per_player >= COURT_ABUNDANCE_THRESHOLD:
            return EARLY_DEPARTURE_GAMES["abundant"]
        else:
            return EARLY_DEPARTURE_GAMES["tight"]

    return None


def parse_signup(signup_text: str) -> Tuple[List[str], List[str]]:
    """Parse signup list and return male and female players."""
    males = []
    females = []

    for name in signup_text.strip().split("\n"):
        name = name.strip()
        if not name or name.startswith("#") or name.startswith("2026"):
            continue
        if "." in name or "," in name:
            name = name.split(".", 1)[-1].split(",", 1)[-1].strip()

        if name in MALE_PLAYERS:
            males.append(name)
        elif name in FEMALE_PLAYERS:
            females.append(name)

    return males, females


def get_court_count(total_players: int) -> int:
    """Determine number of courts based on player count."""
    if total_players <= 12:
        return 2
    return 3


def generate_mixed_doubles_matches(males: List[str], females: List[str]) -> List[Tuple[Tuple[str, str], Tuple[str, str]]]:
    """Generate mixed doubles matches."""
    mixed_males = [m for m in males if m in MIXED_DOUBLES_MALES]

    # 过滤只打女双的球员（但如果女双人数不足 4 人，则允许她们参加混双）
    can_play_womens_doubles = len(females) >= 4
    mixed_females = [f for f in females 
                     if not PLAYER_CONSTRAINTS.get(f, {}).get("only_womens_doubles") 
                     or not can_play_womens_doubles]

    matches = []

    for male_pair in itertools.combinations(mixed_males, 2):
        for female_pair in itertools.combinations(mixed_females, 2):
            match = ((male_pair[0], female_pair[0]), (male_pair[1], female_pair[1]))
            matches.append(match)

    return matches


def generate_mens_doubles_matches(males: List[str]) -> List[Tuple[Tuple[str, str], Tuple[str, str]]]:
    """Generate men's doubles matches."""
    matches = []
    for pair1 in itertools.combinations(males, 2):
        remaining = [m for m in males if m not in pair1]
        if len(remaining) < 2:
            continue
        for pair2 in itertools.combinations(remaining, 2):
            match = (pair1, pair2)
            matches.append(match)
    return matches


def generate_womens_doubles_matches(females: List[str]) -> List[Tuple[Tuple[str, str], Tuple[str, str]]]:
    """Generate women's doubles matches."""
    matches = []
    if len(females) < 4:
        return matches

    for pair1 in itertools.combinations(females, 2):
        remaining = [f for f in females if f not in pair1]
        if len(remaining) < 2:
            continue
        for pair2 in itertools.combinations(remaining, 2):
            match = (pair1, pair2)
            matches.append(match)
    return matches


def get_match_players(match: Tuple[Tuple[str, str], Tuple[str, str]]) -> set:
    """Get all players in a match."""
    players = set()
    for pair in match:
        players.update(pair)
    return players


def get_pair_key(p1: str, p2: str) -> tuple:
    """Get normalized pair key."""
    return tuple(sorted([p1, p2]))


def select_balanced_matches(
    mixed_matches: List, mens_matches: List, womens_matches: List,
    total_matches: int, court_count: int, players: List[str],
    males: List[str], females: List[str]
) -> List[Dict]:
    """Select balanced matches with all constraints."""
    selected = []
    player_games = {p: 0 for p in players}
    partner_games = {pair: 0 for pair in FIXED_PARTNERS}

    def get_constraint(player):
        return PLAYER_CONSTRAINTS.get(player, {})

    def can_player_play(player, current_round_matches, round_num, total_rounds):
        """
        检查球员是否可以在当前轮次上场比赛。

        【最高优先级规则】每个轮次中，每个球员在所有场地（1 号、2 号、3 号）只能出现一次！
        这是排阵的核心原则，必须严格保证。

        【外援球员规则】外援球员最大场次限制为 5 场，内部员工为 6 场，
        优先保障内部员工的场次。
        """
        constraint = get_constraint(player)
        # Use dynamic fixed_games calculation based on court availability
        fixed_games = get_fixed_games_for_player(player, total_matches, len(players))
        max_games = get_max_games_for_player(player)

        # 【固定场次规则】严格执行 fixed_games 限制
        # 有 fixed_games 约束的球员，达到固定场次后禁止参赛（不享受冷场规则例外）
        if fixed_games is not None and player_games[player] >= fixed_games:
            return False

        # 检查是否已达到最大场次限制（外援 5 场，内部 6 场）
        if player_games[player] >= max_games:
            return False

        # 【核心规则】检查球员是否已在当前轮次的其他场地出现过
        for m in current_round_matches:
            if player in get_match_players(m["match"]):
                return False  # 该球员已在本轮次上场比赛，不能再参加

        return True

    def can_add_match(match, current_round_matches, round_num, total_rounds):
        match_type = None
        for m in current_round_matches:
            if m.get("type"):
                match_type = m["type"]
                break

        # 获取当前女性球员总数，判断女双是否可行
        all_females_count = len([p for p in players if p in FEMALE_PLAYERS])
        can_play_womens_doubles = all_females_count >= 4

        for player in get_match_players(match):
            # 检查特殊约束：只打女双的球员不能参加混双（但女双人数不足 4 人时例外）
            constraint = get_constraint(player)
            if constraint.get("only_womens_doubles"):
                # 如果当前比赛不是女双，且女双人数足够 4 人，该球员不能参加
                if match_type and match_type != "女双" and can_play_womens_doubles:
                    return False

            if not can_player_play(player, current_round_matches, round_num, total_rounds):
                return False

        return True

    def get_match_score(match, round_num, is_womens_doubles=False):
        """评分越低越优先被选中。

        评分规则：
        1. 已达到固定场次的球员 = 禁止参赛（巨额加分）
        2. 固定场次要求：未完成时大幅减分（最优先）
        3. 未达到目标场次的球员优先
        4. 外援球员：小幅加分（降低优先级，优先保障内部员工）
        5. 达到最大场次的球员：大幅加分（最后）
        6. 固定搭档组合：减分（优先）
        7. 女双比赛：大幅减分（4 个女生难得，优先安排）
        """
        match_players = get_match_players(match)
        scores = []
        active_games = [g for p, g in player_games.items() if g > 0]
        avg_games = sum(active_games) / len(active_games) if active_games else 0

        TARGET_GAMES = 5  # 目标场次

        for player in match_players:
            constraint = get_constraint(player)
            games = player_games[player]
            # Use dynamic fixed_games calculation
            fixed_games = get_fixed_games_for_player(player, total_matches, len(players))
            max_games = get_max_games_for_player(player)
            score = 0

            # 优先级 1：已达到固定场次的球员 = 禁止参赛（巨额加分）
            if fixed_games is not None and games >= fixed_games:
                score += 10000  # 巨额加分，确保不会被选中

            # 优先级 2：固定场次要求：未完成时大幅减分（最优先）
            if fixed_games is not None:
                remaining = fixed_games - games
                if remaining > 0:
                    score -= remaining * 1000

            # 优先级 3：未达到目标场次的球员优先
            if games < TARGET_GAMES:
                score -= (TARGET_GAMES - games) * 200
            elif games >= max_games:
                # 达到最大场次，大幅加分（靠后）
                score += 2000
            else:
                # 达到 5 场但未到最大场次，小幅加分
                score += (games - TARGET_GAMES) * 100

            # 优先级 4：外援球员：小幅加分（降低优先级，优先保障内部员工）
            if is_guest_player(player):
                score += 50

            scores.append(score)

        pair_a, pair_b = match
        for pair in [pair_a, pair_b]:
            pair_key = get_pair_key(pair[0], pair[1])
            if pair_key in FIXED_PARTNERS:
                weight = FIXED_PARTNERS[pair_key]
                scores.append(-weight * 30)

        # 优先级 8：女双比赛减分（4 个女生难得，优先安排，但不超过目标）
        if is_womens_doubles and womens_used < womens_target:
            remaining_womens = womens_target - womens_used
            round_bonus = 1500 if round_num <= 3 else 600
            scores.append(-remaining_womens * round_bonus)

        return sum(scores) / len(scores) if scores else float('inf')

    mixed_pool = list(mixed_matches)
    mens_pool = list(mens_matches)
    womens_pool = list(womens_matches)

    random.shuffle(mixed_pool)
    random.shuffle(mens_pool)
    random.shuffle(womens_pool)

    # 预设比赛类型比例和目标数量
    # 根据实际女性人数动态调整女双比例
    womens_count = len(females)
    if womens_count >= 4:
        # 4 个女生：女双目标 3-4 场
        mixed_target = int(total_matches * 0.28)
        womens_target = 4
    else:
        mixed_target = int(total_matches * 0.33)
        womens_target = 0

    mens_target = total_matches - mixed_target - womens_target

    mixed_used = 0
    mens_used = 0
    womens_used = 0

    rounds = []
    current_round = []
    total_rounds = (total_matches + court_count - 1) // court_count

    # 比赛类型名称映射
    match_type_name_map = {"mixed": "混双", "mens": "男双", "womens": "女双"}

    def try_add_match(pool, match_type_name, current_count, target_count, allow_fallback=False):
        """尝试添加一场比赛到当前轮次，返回是否成功。
        
        Args:
            pool: 比赛池
            match_type_name: 比赛类型 (mixed/mens/womens)
            current_count: 当前已使用该类型的次数
            target_count: 目标使用次数
            allow_fallback: 是否允许 fallback（混双/女双位置用男双代替）
        """
        nonlocal current_round, mixed_used, mens_used, womens_used

        # 如果已达到目标数量，跳过（女双严格限制，其他类型允许稍微超出）
        if match_type_name == "womens" and current_count >= womens_target:
            return False, current_count
        if match_type_name != "womens" and current_count >= target_count + 2:
            return False, current_count

        current_round_num = len(rounds) + 1
        best_match = None
        best_score = float('inf')

        # 获取当前女性球员总数，判断女双是否可行
        all_females_count = len([p for p in players if p in FEMALE_PLAYERS])
        can_play_womens_doubles = all_females_count >= 4

        # Fallback 逻辑：如果允许 fallback 且原池子为空或无法提供比赛，从男双池借
        if allow_fallback and (not pool or (match_type_name == "mixed" and len(females) < 2)):
            # 从男双池借一场比赛来打这个位置
            if mens_pool:
                for match in mens_pool:
                    if can_add_match(match, current_round, current_round_num, total_rounds):
                        score = get_match_score(match, current_round_num, is_womens_doubles=False)
                        if score < best_score:
                            best_score = score
                            best_match = match

        # 如果没有找到 fallback 比赛，尝试原池子
        if not best_match and pool:
            for match in pool:
                # 检查是否有球员不能参加该类型的比赛
                can_play = True
                for player in get_match_players(match):
                    constraint = get_constraint(player)
                    if constraint.get("only_womens_doubles") and match_type_name != "womens" and can_play_womens_doubles:
                        can_play = False
                        break

                if not can_play:
                    continue

                if not can_add_match(match, current_round, current_round_num, total_rounds):
                    continue
                is_wd = (match_type_name == "womens")
                score = get_match_score(match, current_round_num, is_womens_doubles=is_wd)
                if score < best_score:
                    best_score = score
                    best_match = match

        if best_match:
            # 确定实际显示的类型
            is_fallback = best_match in mens_pool and match_type_name != "mens"
            if is_fallback:
                display_type = f"{match_type_name_map[match_type_name]} (男代)"
            else:
                display_type = match_type_name_map[match_type_name]
            
            current_round.append({
                "type": display_type,
                "match": best_match,
                "fallback": is_fallback
            })
            for player in get_match_players(best_match):
                player_games[player] += 1
            pair_a, pair_b = best_match
            for pair in [pair_a, pair_b]:
                pair_key = get_pair_key(pair[0], pair[1])
                if pair_key in partner_games:
                    partner_games[pair_key] += 1
            
            # 从对应的池子移除
            if best_match in pool:
                pool.remove(best_match)
            elif best_match in mens_pool:
                mens_pool.remove(best_match)
            
            # 更新计数（fallback 也算作原类型的计数）
            if match_type_name == "mixed":
                mixed_used += 1
            elif match_type_name == "mens":
                mens_used += 1
            else:
                womens_used += 1

            return True, current_count + 1
        return False, current_count

    # 多轮填充：每轮尝试填满所有场地
    # 策略：女双始终最优先，直到达到目标
    # 放宽女双目标限制，在未达到目标前不计上限
    type_priorities_default = [
        ("womens", lambda: womens_used, lambda: womens_target + 10, womens_pool),  # 女双最优先（大幅放宽上限）
        ("mixed", lambda: mixed_used, lambda: mixed_target + 10, mixed_pool),  # 混双次优先（也放宽）
        ("mens", lambda: mens_used, lambda: mens_target + 10, mens_pool),  # 男双最后
    ]

    # 所有轮次都用同样的优先级（女双始终优先）
    type_priorities_mixed_first = type_priorities_default

    for round_num in range(total_rounds):
        current_round = []

        # 交替优先级：奇数轮女双优先，偶数轮混双优先
        if round_num % 2 == 0:
            type_priorities = type_priorities_default  # 女双优先（0, 2, 4...）
        else:
            type_priorities = type_priorities_mixed_first  # 混双优先（1, 3, 5...）

        # 尝试填满当前轮次的所有场地
        for court_slot in range(court_count):
            added = False

            # 按优先级尝试每种比赛类型
            for match_type, get_used, get_target, pool in type_priorities:
                if not pool:
                    continue
                success, new_count = try_add_match(pool, match_type, get_used(), get_target())
                if success:
                    added = True
                    # 更新计数
                    if match_type == "mixed":
                        mixed_used = new_count
                    elif match_type == "womens":
                        womens_used = new_count
                    else:
                        mens_used = new_count
                    break

            # 如果优先级类型都无法添加，尝试所有类型（放宽限制）
            if not added:
                for match_type, get_used, get_target, pool in type_priorities:
                    if not pool:
                        continue
                    # 放宽限制：不检查目标数量
                    success, new_count = try_add_match(pool, match_type, 0, 999)
                    if success:
                        added = True
                        if match_type == "mixed":
                            mixed_used = new_count
                        elif match_type == "womens":
                            womens_used = new_count
                        else:
                            mens_used = new_count
                        break

            # 如果还是无法添加，尝试 fallback 模式（用男双填充场地）
            if not added:
                # 尝试用男双代替混双：从男双池找一场可用的比赛
                if mens_pool:
                    for match in list(mens_pool):
                        if can_add_match(match, current_round, len(rounds) + 1, total_rounds):
                            current_round.append({
                                "type": "混双 (男代)",
                                "match": match,
                                "fallback": True
                            })
                            for player in get_match_players(match):
                                player_games[player] += 1
                            pair_a, pair_b = match
                            for pair in [pair_a, pair_b]:
                                pair_key = get_pair_key(pair[0], pair[1])
                                if pair_key in partner_games:
                                    partner_games[pair_key] += 1
                            mens_pool.remove(match)
                            mens_used += 1
                            added = True
                            break

            # 如果还是无法添加，尝试临时组合：从本轮未上场的人中选 4 个打第 3 场地
            if not added:
                # 找出本轮已上场的球员
                played_players = set()
                for m in current_round:
                    played_players.update(get_match_players(m["match"]))
                
                # 候选人 = 所有球员 - 已上场球员 - 不能打的球员（fixed_games 已满）
                candidates = []
                for p in players:
                    if p in played_players:
                        continue
                    # Use dynamic fixed_games calculation
                    fixed_games = get_fixed_games_for_player(p, total_matches, len(players))
                    if fixed_games is not None and player_games[p] >= fixed_games:
                        continue
                    max_games = get_max_games_for_player(p)
                    if player_games[p] >= max_games:
                        continue
                    candidates.append(p)
                
                # 如果有至少 4 个候选人，随机选 4 个组成一场比赛
                if len(candidates) >= 4:
                    selected = random.sample(candidates, 4)
                    # 尽量按性别配对
                    females_in_selected = [p for p in selected if p in FEMALE_PLAYERS]
                    males_in_selected = [p for p in selected if p not in FEMALE_PLAYERS]
                    
                    if len(females_in_selected) >= 2 and len(males_in_selected) >= 2:
                        # 2 女 2 男：女双
                        pair1 = (females_in_selected[0], females_in_selected[1])
                        pair2 = (males_in_selected[0], males_in_selected[1])
                        match_type = "女双 (乱打)"
                    elif len(males_in_selected) >= 4:
                        # 4 男：男双
                        pair1 = (males_in_selected[0], males_in_selected[1])
                        pair2 = (males_in_selected[2], males_in_selected[3])
                        match_type = "男双 (乱打)"
                    elif len(males_in_selected) >= 3:
                        # 3 男 1 女：男双（2 男 vs 1 男 + 女打男双位置）
                        pair1 = (males_in_selected[0], males_in_selected[1])
                        pair2 = (males_in_selected[2], females_in_selected[0])
                        match_type = "男双 (乱打)"
                    elif len(males_in_selected) == 2 and len(females_in_selected) == 2:
                        # 2 男 2 女：女双或混双
                        pair1 = (females_in_selected[0], females_in_selected[1])
                        pair2 = (males_in_selected[0], males_in_selected[1])
                        match_type = "女双 (乱打)"
                    else:
                        # 默认：按 selected 顺序配对
                        pair1 = (selected[0], selected[1])
                        pair2 = (selected[2], selected[3])
                        match_type = "男双 (乱打)"
                    
                    match = (pair1, pair2)
                    current_round.append({
                        "type": match_type,
                        "match": match,
                        "fallback": True
                    })
                    for player in get_match_players(match):
                        player_games[player] += 1
                    added = True

            if not added:
                break

        if current_round:
            rounds.append(current_round)
            
            # 【防止冷场规则】更新球员的轮空历史
            # 找出本轮上场的球员
            played_players = set()
            for m in current_round:
                played_players.update(get_match_players(m["match"]))

    # 【核心规则检查】验证没有球员在同一轮次重复出现

    # 直接使用已构建的轮次结构
    # 为每个比赛添加轮次和场地信息
    selected = []
    for round_idx, round_matches in enumerate(rounds):
        # 【核心规则检查】每个轮次中，每个球员只能出现一次
        round_players = set()
        for match_info in round_matches:
            players_in_match = get_match_players(match_info["match"])
            for p in players_in_match:
                if p in round_players:
                    raise ValueError(f"【核心规则违规】第{round_idx + 1}轮：球员 {p} 在同一轮次重复出现！")
                round_players.add(p)
        
        for court_idx, match_info in enumerate(round_matches):
            match_info["court"] = court_idx + 1
            match_info["round"] = round_idx + 1
            selected.append(match_info)

    return selected


def create_lineup_excel(matches: List[Dict], court_count: int, output_path: str, player_stats: Dict = None, activity_date: str = None):
    """Create Excel file with lineup schedule."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对阵表"

    font_title = Font(name="微软雅黑", size=18, bold=True)
    font_header = Font(name="微软雅黑", size=11, bold=True)
    font_content = Font(name="微软雅黑", size=10)
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws.merge_cells("A1:G1")
    title = "科技球队日常训练活动 - 对阵表"
    if activity_date:
        title = f"科技球队日常训练活动 - 对阵表（{activity_date}）"
    ws["A1"] = title
    ws["A1"].font = font_title
    ws["A1"].alignment = alignment_center
    ws["A1"].border = border

    ws.merge_cells("A2:G2")
    ws["A2"] = f"场地数：{court_count}个 | 时长：2 小时 | 赛制：15 分/局，2 局 | 项目：男双、女双、混双"
    ws["A2"].font = Font(name="微软雅黑", size=10)
    ws["A2"].alignment = alignment_center

    headers = ["轮次", "场地", "类型", "对阵 A", "比分 A", "比分 B", "对阵 B"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = border
        cell.fill = header_fill

    for row_idx, match_info in enumerate(matches, 4):
        match_type = match_info["type"]
        match = match_info["match"]
        court = match_info["court"]
        round_num = match_info.get("round", 1)  # 直接使用已存储的轮次值

        ws.cell(row=row_idx, column=1, value=round_num).font = font_content
        ws.cell(row=row_idx, column=2, value=f"{court}号").font = font_content
        ws.cell(row=row_idx, column=3, value=match_type).font = font_content
        ws.cell(row=row_idx, column=4, value="/".join(match[0])).font = font_content
        ws.cell(row=row_idx, column=5, value="").font = font_content
        ws.cell(row=row_idx, column=6, value="").font = font_content
        ws.cell(row=row_idx, column=7, value="/".join(match[1])).font = font_content

        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).alignment = alignment_center
            ws.cell(row=row_idx, column=col).border = border

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 22

    max_row = len(matches) + 3
    for row in ws.iter_rows(min_row=1, max_row=max_row if max_row > 3 else 20):
        ws.row_dimensions[row[0].row].height = 32

    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "landscape"
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    if player_stats:
        ws_stats = wb.create_sheet(title="球员统计")
        font_title = Font(name="微软雅黑", size=14, bold=True)
        font_header = Font(name="微软雅黑", size=11, bold=True)
        font_content = Font(name="微软雅黑", size=10)
        alignment_center = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        ws_stats.merge_cells("A1:E1")
        ws_stats["A1"] = "球员参赛场次统计"
        ws_stats["A1"].font = font_title
        ws_stats["A1"].alignment = alignment_center
        ws_stats["A1"].border = border

        headers = ["姓名", "总场次", "男双", "女双", "混双"]
        for col, header in enumerate(headers, 1):
            cell = ws_stats.cell(row=2, column=col, value=header)
            cell.font = font_header
            cell.alignment = alignment_center
            cell.border = border
            cell.fill = header_fill

        row = 3
        male_players = [p for p in player_stats.keys() if p in MALE_PLAYERS]
        female_players = [p for p in player_stats.keys() if p in FEMALE_PLAYERS]

        for player in sorted(male_players):
            stats = player_stats[player]
            ws_stats.cell(row=row, column=1, value=player).font = font_content
            ws_stats.cell(row=row, column=2, value=stats["total"]).font = font_content
            ws_stats.cell(row=row, column=3, value=stats.get("男双", 0)).font = font_content
            ws_stats.cell(row=row, column=4, value=stats.get("女双", 0)).font = font_content
            ws_stats.cell(row=row, column=5, value=stats.get("混双", 0)).font = font_content
            for col in range(1, 6):
                ws_stats.cell(row=row, column=col).alignment = alignment_center
                ws_stats.cell(row=row, column=col).border = border
            row += 1

        for player in sorted(female_players):
            stats = player_stats[player]
            ws_stats.cell(row=row, column=1, value=player).font = font_content
            ws_stats.cell(row=row, column=2, value=stats["total"]).font = font_content
            ws_stats.cell(row=row, column=3, value=stats.get("男双", 0)).font = font_content
            ws_stats.cell(row=row, column=4, value=stats.get("女双", 0)).font = font_content
            ws_stats.cell(row=row, column=5, value=stats.get("混双", 0)).font = font_content
            for col in range(1, 6):
                ws_stats.cell(row=row, column=col).alignment = alignment_center
                ws_stats.cell(row=row, column=col).border = border
            row += 1

        ws_stats.column_dimensions["A"].width = 12
        ws_stats.column_dimensions["B"].width = 10
        ws_stats.column_dimensions["C"].width = 10
        ws_stats.column_dimensions["D"].width = 10
        ws_stats.column_dimensions["E"].width = 10

    wb.save(output_path)
    print(f"对阵表已生成：{output_path}")


def main():
    # Try multiple paths for flexibility
    possible_paths = [
        "微信接龙.txt",  # When running from 排阵 directory
        "排阵/微信接龙.txt",  # When running from project root
    ]

    signup_text = None
    for signup_file in possible_paths:
        try:
            with open(signup_file, "r", encoding="utf-8") as f:
                signup_text = f.read()
            break
        except FileNotFoundError:
            continue

    if signup_text is None:
        print(f"错误：找不到报名文件，请在以下路径之一创建文件:")
        for path in possible_paths:
            print(f"  - {path}")
        return

    # Parse activity date
    activity_date = parse_activity_date(signup_text)
    
    males, females = parse_signup(signup_text)
    all_players = males + females
    total_players = len(all_players)

    print(f"报名男性 ({len(males)}人): {', '.join(males)}")
    print(f"报名女性 ({len(females)}人): {', '.join(females)}")
    print(f"总人数：{total_players}人")
    if activity_date:
        print(f"活动日期：{activity_date}")

    court_count = get_court_count(total_players)
    print(f"场地数量：{court_count}个")

    mixed_matches = generate_mixed_doubles_matches(males, females)
    mens_matches = generate_mens_doubles_matches(males)
    womens_matches = generate_womens_doubles_matches(females)

    # Calculate actual available player-games
    court_count_temp = get_court_count(len(all_players))
    target_matches_temp = MATCHES_PER_COURT * court_count_temp
    total_available_games = 0
    for p in all_players:
        constraint = PLAYER_CONSTRAINTS.get(p, {})
        # Use dynamic fixed_games calculation based on court availability
        fixed_games = get_fixed_games_for_player(p, target_matches_temp, len(all_players))
        if fixed_games is not None:
            total_available_games += fixed_games
        else:
            total_available_games += get_max_games_for_player(p)

    # Each match requires 4 player-games
    max_possible_matches = total_available_games // 4
    target_matches = MATCHES_PER_COURT * court_count
    total_matches = min(max_possible_matches, target_matches)

    print(f"\n配置参数:")
    print(f"  - 每场地比赛数：{MATCHES_PER_COURT}场")
    print(f"  - 可用总人次：{total_available_games}")
    print(f"  - 理论最大比赛数：{max_possible_matches}场")
    print(f"  - 目标比赛数：{target_matches}场")
    print(f"  - 实际比赛数：{total_matches}场（受限于球员场次）")
    print(f"  - 内部员工最大场次：{MAX_GAMES_INTERNAL}场")
    print(f"  - 外援球员最大场次：{MAX_GAMES_GUEST}场（优先保障内部员工）")

    selected_matches = select_balanced_matches(
        mixed_matches, mens_matches, womens_matches, total_matches, court_count, all_players, males, females
    )

    from collections import Counter
    match_type_counts = Counter()
    for m in selected_matches:
        match_type_counts[m["type"]] += 1

    print(f"\n比赛类型分布:")
    for t, c in sorted(match_type_counts.items()):
        print(f"  - {t}: {c}场")

    # 【轮次详情】输出每个轮次每个场地的球员，人工确认无重复
    print(f"\n【轮次详情】")
    rounds_dict = {}
    for m in selected_matches:
        round_num = m.get("round", 1)
        if round_num not in rounds_dict:
            rounds_dict[round_num] = []
        rounds_dict[round_num].append(m)
    
    for round_num in sorted(rounds_dict.keys()):
        round_matches = rounds_dict[round_num]
        court_count_round = len(round_matches)  # 本轮实际使用的场地数
        print(f"\n第{round_num}轮 ({court_count_round}个场地):")
        round_players = set()
        for m in round_matches:
            court = m["court"]
            match_type = m["type"]
            team_a = "/".join(m["match"][0])
            team_b = "/".join(m["match"][1])
            players_in_match = set(get_match_players(m["match"]))
            print(f"  {court}号场地 [{match_type}]: {team_a} VS {team_b}")
            round_players.update(players_in_match)
        print(f"  本轮球员 ({len(round_players)}人): {', '.join(sorted(round_players))}")
        # 计算轮空球员（排除提前走的人）
        round_bye_players = [p for p in all_players if p not in round_players and not PLAYER_CONSTRAINTS.get(p, {}).get("early_departure")]
        if round_bye_players:
            print(f"  本轮轮空 ({len(round_bye_players)}人): {', '.join(sorted(round_bye_players))}")
    
    print("\n✓ 核心规则检查通过：无球员在同一轮次重复出现")

    print(f"\n球员参赛场次统计:")
    player_games = {}
    player_type_games = {}
    for m in selected_matches:
        match_type = m["type"]
        # 处理 fallback 类型：提取基础类型（男双/女双/混双）
        base_type = match_type.split(" ")[0]  # "男双 (临时)" -> "男双", "混双 (男代)" -> "混双"
        for player in get_match_players(m["match"]):
            player_games[player] = player_games.get(player, 0) + 1
            if player not in player_type_games:
                player_type_games[player] = {"男双": 0, "女双": 0, "混双": 0}
            player_type_games[player][base_type] += 1

    for player in sorted(all_players):
        games = player_games.get(player, 0)
        # Use dynamic fixed_games calculation for display
        fixed_games = get_fixed_games_for_player(player, total_matches, len(all_players))
        note = ""
        if fixed_games is not None:
            if games == fixed_games:
                note = f" (已完成{fixed_games}场)"
            elif games < fixed_games:
                note = f" (目标{fixed_games}场，还差{fixed_games - games}场)"
        if PLAYER_CONSTRAINTS.get(player, {}).get("early_departure"):
            note += " [需提前离场]"
        type_stats = player_type_games.get(player, {"男双": 0, "女双": 0, "混双": 0})
        type_str = f"(男{type_stats['男双']}/女{type_stats['女双']}/混{type_stats['混双']})"
        print(f"  - {player}: {games}场 {type_str}{note}")

    player_stats = {}
    for player in all_players:
        type_stats = player_type_games.get(player, {"男双": 0, "女双": 0, "混双": 0})
        player_stats[player] = {
            "total": player_games.get(player, 0),
            "男双": type_stats["男双"],
            "女双": type_stats["女双"],
            "混双": type_stats["混双"]
        }

    # Try multiple paths for flexibility
    possible_output_paths = [
        "对阵表.xlsx",  # When running from 排阵 directory
        "排阵/对阵表.xlsx",  # When running from project root
    ]
    
    output_path = possible_output_paths[0]
    for path in possible_output_paths:
        try:
            # Test if we can write to this path
            with open(path, 'wb') as f:
                f.close()
            import os
            os.remove(path)  # Remove test file
            output_path = path
            break
        except (IOError, OSError):
            continue
    
    create_lineup_excel(
        selected_matches, court_count, output_path, player_stats,
        activity_date=activity_date
)


if __name__ == "__main__":
    main()
