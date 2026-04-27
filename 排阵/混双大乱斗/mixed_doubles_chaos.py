#!/usr/bin/env python3
"""
混双大乱斗排阵系统
A组男队员和B组女队员依次组合对战
支持2个场地同时进行，确保同一轮次球员不重复
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import List, Tuple, Dict
import itertools


def generate_mixed_doubles_matches(
    group_a_males: List[str],
    group_b_females: List[str],
    court_count: int = 2,
    mode: str = "fair",
    priority_players: List[str] = None
) -> List[Dict]:
    """
    生成混双大乱斗对阵表
    
    Args:
        group_a_males: A组男队员列表
        group_b_females: B组女队员列表
        court_count: 场地数量（默认2）
        mode: 排阵模式
            - "fair": 公平模式，每个男队员与每个女队员组合一次（推荐）
            - "max": 最大场次模式，尽可能多的比赛
        priority_players: 需要优先排阵的球员列表（如需要提前回家的球员）
    
    Returns:
        比赛列表，每个元素包含：
        - round: 轮次
        - court: 场地号
        - type: 比赛类型
        - match: ((男1, 女1), (男2, 女2))
    """
    matches = []
    
    if mode == "fair":
        return _generate_fair_mode(group_a_males, group_b_females, court_count, priority_players)
    else:
        return _generate_max_mode(group_a_males, group_b_females, court_count)


def _generate_fair_mode(
    group_a_males: List[str],
    group_b_females: List[str],
    court_count: int = 2,
    priority_players: List[str] = None
) -> List[Dict]:
    """
    公平模式：严格按照优先级生成赛程
    
    优先级：
    1. 【必须】每个人场次一样（6场）
    2. 【必须】休息1轮后必须上场
    3. 【可选】每个男队员尽量和不同女队员搭配
    4. 【可选】尽量每轮2个场地并发
    5. 【特殊】指定球员优先排阵（不休息，排在前面轮次）
    
    Args:
        priority_players: 需要优先排阵的球员列表（如需要提前回家的球员）
    """
    import random
    
    if priority_players is None:
        priority_players = []
    
    males = group_a_males[:]
    females = group_b_females[:]
    n = len(males)
    target_games = n  # 每人应该打6场
    
    matches = []
    pair_used = set()  # 已使用的(男,女)配对
    player_games = {p: 0 for p in males + females}  # 每人已打场次
    last_round_players = set()  # 上一轮上场的球员
    rested_last_round = set()  # 上一轮休息的球员
    
    current_round = 1
    max_rounds = 20
    
    while current_round <= max_rounds:
        used_in_round = set()
        round_matches = []
        
        # 找出所有可以安排的比赛候选
        candidates = []
        for m1_idx, male1 in enumerate(males):
            for f1_idx, female1 in enumerate(females):
                pair1 = (male1, female1)
                if pair1 in pair_used:
                    continue
                
                for m2_idx, male2 in enumerate(males):
                    if m2_idx == m1_idx:
                        continue
                    for f2_idx, female2 in enumerate(females):
                        if f2_idx == f1_idx:
                            continue
                        
                        pair2 = (male2, female2)
                        if pair2 in pair_used:
                            continue
                        
                        match = ((male1, female1), (male2, female2))
                        match_players = {male1, male2, female1, female2}
                        
                        # 检查是否与本轮已安排的球员冲突
                        if match_players & used_in_round:
                            continue
                        
                        # 计算优先级分数
                        priority = 0
                        
                        # 优先级5【特殊】：包含优先排阵球员（最高优先级）
                        priority_in_match = match_players & set(priority_players)
                        can_play_priority = [p for p in priority_in_match if player_games[p] < target_games]
                        priority += len(can_play_priority) * 50000
                        
                        # 优先级2【必须】：包含上一轮休息且未满6场的球员（高优先级）
                        rested_in_match = match_players & rested_last_round
                        can_play = [p for p in rested_in_match if player_games[p] < target_games]
                        priority += len(can_play) * 10000
                        
                        # 优先级1【必须】：优先安排场次少的球员
                        min_games = min(player_games[p] for p in match_players)
                        max_games = max(player_games[p] for p in match_players)
                        # 如果有人已经打满6场，这场比赛不应该被优先考虑
                        if max_games >= target_games:
                            priority -= 5000
                        priority += (target_games - min_games) * 100
                        
                        # 优先级3：尽量使用未使用过的配对
                        new_pairs = sum(1 for p in [pair1, pair2] if p not in pair_used)
                        priority += new_pairs * 10
                        
                        candidates.append((priority, match, match_players, pair1, pair2))
        
        if not candidates:
            break
        
        # 按优先级排序
        candidates.sort(key=lambda x: x[0], reverse=True)
        
        # 尝试安排最多2场比赛
        for priority, match, match_players, pair1, pair2 in candidates:
            if len(round_matches) >= 2:
                break
            
            # 再次检查是否冲突（因为used_in_round可能在循环中更新）
            if match_players & used_in_round:
                continue
            
            round_matches.append(match)
            used_in_round.update(match_players)
            pair_used.add(pair1)
            pair_used.add(pair2)
            
            for p in match_players:
                player_games[p] += 1
        
        if round_matches:
            for court_idx, match in enumerate(round_matches):
                matches.append({
                    "round": current_round,
                    "court": court_idx + 1,
                    "type": "混双",
                    "match": match
                })
            
            # 更新状态
            rested_last_round = set(males + females) - used_in_round
            last_round_players = used_in_round.copy()
            current_round += 1
        else:
            break
        
        # 检查是否所有人都打了6场
        if all(g == target_games for g in player_games.values()):
            break
    
    # 验证是否满足要求
    if not all(g == target_games for g in player_games.values()):
        print(f"⚠️  警告：未能让所有人都打{target_games}场")
        for p, g in player_games.items():
            if g != target_games:
                print(f"  {p}: {g}场")
    
    return matches


def _generate_max_mode(
    group_a_males: List[str],
    group_b_females: List[str],
    court_count: int = 2
) -> List[Dict]:
    """
    最大场次模式：尽可能多的比赛（原有逻辑）
    """
    import random
    
    matches = []
    
    # 生成所有可能的混双组合
    all_pairs = [(male, female) for male in group_a_males for female in group_b_females]
    
    if len(all_pairs) < 2:
        print("⚠️  配对数量不足，无法进行比赛")
        return matches
    
    # 默认最多15轮
    max_rounds = min(len(group_a_males) * len(group_b_females), 15)
    
    # 按轮次安排比赛，每轮使用不同的球员
    used_in_round = set()  # 当前轮次已使用的球员
    current_round = 1
    round_matches = []
    used_combinations = set()  # 记录已经使用过的对决组合
    
    # 遍历所有可能的对决组合
    pair_combinations = list(itertools.combinations(all_pairs, 2))
    
    # 打乱顺序以增加随机性
    random.shuffle(pair_combinations)
    
    for pair1, pair2 in pair_combinations:
        # 如果已经达到最大轮数，停止
        if current_round > max_rounds:
            break
        
        # 检查这4个球员是否在当前轮次已经出现过
        players_needed = set(pair1 + pair2)
        
        # 关键检查：确保两个配对没有共同的男队员或女队员
        male1, female1 = pair1
        male2, female2 = pair2
        
        # 不能有相同的男队员或女队员
        if male1 == male2 or female1 == female2:
            continue
        
        # 检查是否有重复球员
        has_overlap = bool(players_needed & used_in_round)
        
        # 检查这个组合是否已经用过
        combo_key = tuple(sorted([tuple(sorted(pair1)), tuple(sorted(pair2))]))
        already_used = combo_key in used_combinations
        
        if has_overlap or already_used:
            continue
        
        # 添加这场比赛
        round_matches.append((pair1, pair2))
        used_in_round.update(players_needed)
        used_combinations.add(combo_key)
        
        # 如果达到场地数量限制，开始新轮次
        if len(round_matches) >= court_count:
            for i, match in enumerate(round_matches):
                matches.append({
                    "round": current_round,
                    "court": i + 1,
                    "type": "混双",
                    "match": match
                })
            current_round += 1
            used_in_round.clear()
            round_matches = []
    
    # 处理剩余的比赛
    if round_matches and current_round <= max_rounds:
        for i, match in enumerate(round_matches):
            matches.append({
                "round": current_round,
                "court": i + 1,
                "type": "混双",
                "match": match
            })
    
    return matches


def create_excel(
    matches: List[Dict],
    group_a_males: List[str],
    group_b_females: List[str],
    output_path: str = "混双大乱斗_对阵表.xlsx"
):
    """
    创建Excel对阵表
    
    Args:
        matches: 比赛列表
        group_a_males: A组男队员列表
        group_b_females: B组女队员列表
        output_path: 输出文件路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对阵表"
    
    # 样式定义
    font_title = Font(name="微软雅黑", size=18, bold=True)
    font_header = Font(name="微软雅黑", size=11, bold=True)
    font_content = Font(name="微软雅黑", size=10)
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # 标题行
    ws.merge_cells("A1:G1")
    ws["A1"] = "混双大乱斗 - 对阵表"
    ws["A1"].font = font_title
    ws["A1"].alignment = alignment_center
    ws["A1"].border = border
    
    # 配置行
    ws.merge_cells("A2:G2")
    ws["A2"] = f"A组男队员 ({len(group_a_males)}人) vs B组女队员 ({len(group_b_females)}人) | 场地数：2个 | 赛制：混双对决"
    ws["A2"].font = Font(name="微软雅黑", size=10)
    ws["A2"].alignment = alignment_center
    
    # 表头
    headers = ["轮次", "场地", "类型", "对阵 A (男/女)", "比分 A", "比分 B", "对阵 B (男/女)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = border
        cell.fill = header_fill
    
    # 比赛数据行
    for row_idx, match_info in enumerate(matches, 4):
        match = match_info["match"]
        court = match_info["court"]
        round_num = match_info["round"]
        
        # 对阵A: 男/女
        pair_a = f"{match[0][0]}/{match[0][1]}"
        # 对阵B: 男/女
        pair_b = f"{match[1][0]}/{match[1][1]}"
        
        ws.cell(row=row_idx, column=1, value=round_num).font = font_content
        ws.cell(row=row_idx, column=2, value=f"{court}号").font = font_content
        ws.cell(row=row_idx, column=3, value="混双").font = font_content
        ws.cell(row=row_idx, column=4, value=pair_a).font = font_content
        ws.cell(row=row_idx, column=5, value="").font = font_content
        ws.cell(row=row_idx, column=6, value="").font = font_content
        ws.cell(row=row_idx, column=7, value=pair_b).font = font_content
        
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).alignment = alignment_center
            ws.cell(row=row_idx, column=col).border = border
    
    # 列宽设置
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 22
    
    # 行高设置
    max_row = len(matches) + 3
    for row in ws.iter_rows(min_row=1, max_row=max_row if max_row > 3 else 20):
        ws.row_dimensions[row[0].row].height = 32
    
    # 页面设置
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "landscape"
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    
    wb.save(output_path)
    print(f"✅ 对阵表已生成：{output_path}")


def print_schedule_summary(matches: List[Dict]):
    """打印赛程摘要"""
    print(f"\n{'='*60}")
    print(f"混双大乱斗赛程摘要")
    print(f"{'='*60}")
    
    print(f"\n总比赛场次: {len(matches)}场")
    
    # 按轮次分组
    rounds_dict = {}
    for m in matches:
        round_num = m["round"]
        if round_num not in rounds_dict:
            rounds_dict[round_num] = []
        rounds_dict[round_num].append(m)
    
    print(f"\n共 {len(rounds_dict)} 轮比赛:\n")
    
    for round_num in sorted(rounds_dict.keys()):
        round_matches = rounds_dict[round_num]
        print(f"第{round_num}轮 ({len(round_matches)}个场地同时进行):")
        
        round_players = set()
        for m in round_matches:
            court = m["court"]
            match = m["match"]
            
            team_a = f"{match[0][0]}/{match[0][1]}"
            team_b = f"{match[1][0]}/{match[1][1]}"
            
            players_in_match = set(match[0] + match[1])
            round_players.update(players_in_match)
            
            print(f"  {court}号场地: {team_a} VS {team_b}")
        
        print(f"  本轮球员 ({len(round_players)}人): {', '.join(sorted(round_players))}\n")
    
    # 统计每个球员的参赛次数
    player_games = {}
    for m in matches:
        match = m["match"]
        for player in match[0] + match[1]:
            player_games[player] = player_games.get(player, 0) + 1
    
    print(f"\n球员参赛统计:")
    for player, games in sorted(player_games.items(), key=lambda x: -x[1]):
        print(f"  - {player}: {games}场")
    
    # 验证：检查是否有球员在同一轮次重复出现
    print(f"\n{'='*60}")
    print(f"核心规则验证:")
    print(f"{'='*60}")
    
    validation_passed = True
    for round_num, round_matches in rounds_dict.items():
        round_players = set()
        for m in round_matches:
            players = set(m["match"][0] + m["match"][1])
            overlap = round_players & players
            if overlap:
                print(f"❌ 第{round_num}轮发现重复球员: {overlap}")
                validation_passed = False
            round_players.update(players)
    
    if validation_passed:
        print(f"✅ 所有轮次均无球员重复，验证通过！")
    
    print(f"{'='*60}\n")


def main():
    """主函数"""
    # 示例数据
    group_a_males = ["林锋", "王小波", "陈顺星", "罗琴荩", "苏大哲", "黄冬青"]
    group_b_females = ["田茜", "李祺祺", "高洁", "滕菲", "崔倩男", "李杏芝"]
    
    print("=" * 60)
    print("混双大乱斗排阵系统")
    print("=" * 60)
    print(f"\nA组男队员 ({len(group_a_males)}人): {', '.join(group_a_males)}")
    print(f"B组女队员 ({len(group_b_females)}人): {', '.join(group_b_females)}")
    
    # 生成对阵表（崔倩男优先排阵）
    print(f"\n正在生成对阵表（2个场地并发，崔倩男优先排阵）...")
    matches = generate_mixed_doubles_matches(
        group_a_males=group_a_males,
        group_b_females=group_b_females,
        court_count=2,
        mode="fair",  # 公平模式：每人6场
        priority_players=["崔倩男"]  # 崔倩男优先排阵，不休息
    )
    
    if not matches:
        print("❌ 未能生成比赛")
        return
    
    # 打印摘要
    print_schedule_summary(matches)
    
    # 生成Excel
    create_excel(
        matches=matches,
        group_a_males=group_a_males,
        group_b_females=group_b_females,
        output_path="混双大乱斗_对阵表.xlsx"
    )


if __name__ == "__main__":
    main()
