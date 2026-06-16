import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import random
import datetime

src_wb = openpyxl.load_workbook('签到表.xlsx')
src_ws = src_wb.active

people = []
for row in src_ws.iter_rows(min_row=3, max_row=src_ws.max_row, max_col=4, values_only=True):
    name, dept, sign, ratio = row
    if name:
        people.append({'name': name, 'dept': dept, 'ratio': ratio or 0.1})

print(f"Total personnel: {len(people)}")

def get_workday_mondays(start_date, end_date):
    mondays = []
    current = start_date
    while current <= end_date:
        if current.weekday() == 0:
            mondays.append(current)
        current += datetime.timedelta(days=1)
    return mondays

start = datetime.date(2026, 1, 1)
end = datetime.date(2026, 5, 31)
all_mondays = get_workday_mondays(start, end)

print(f"\nTotal Mondays from {start} to {end}: {len(all_mondays)}")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def select_attendees(people_list, seed):
    random.seed(seed)
    result = []
    for p in people_list:
        if random.random() < p['ratio']:
            result.append(p)
    random.shuffle(result)
    return result

for idx, target_date in enumerate(all_mondays):
    date_str = target_date.strftime('%Y-%m-%d')
    filename = f"报名表_{date_str}.xlsx"

    new_wb = openpyxl.Workbook()
    ws = new_wb.active
    ws.title = "Sheet1"

    ws.merge_cells('A1:C1')
    title = ws['A1']
    title.value = f"平安科技羽毛球队训练 {date_str}"
    title.font = Font(name='宋体', size=18, bold=False)
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    headers = ['姓名', '部门', '签名']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name='宋体', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 32

    seed = target_date.toordinal()
    attendees = select_attendees(people, seed)
    print(f"\n{date_str}: {len(attendees)} attendees")
    for a in attendees:
        print(f"  {a['name']}")

    for row_idx, person in enumerate(attendees, 3):
        ws.cell(row=row_idx, column=1, value=person['name']).font = Font(name='宋体', size=12)
        ws.cell(row=row_idx, column=2, value=person['dept']).font = Font(name='宋体', size=11)
        ws.cell(row=row_idx, column=3, value=None).font = Font(name='宋体', size=14)
        ws.row_dimensions[row_idx].height = 32

    ws['M27'] = "；。'"

    ws.column_dimensions['A'].width = 11.06
    ws.column_dimensions['B'].width = 58.85
    ws.column_dimensions['C'].width = 18.42
    ws.column_dimensions['M'].width = 13.0

    last_data_row = 2 + len(attendees)
    for row_idx in range(2, last_data_row + 1):
        for col_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    new_wb.save(filename)
    print(f"Saved: {filename}")

print("\nDone!")